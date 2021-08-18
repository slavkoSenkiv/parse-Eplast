import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

browser = webdriver.Chrome(executable_path=r'/Users/ysenkiv/Code/chromedriver')

# entering main page
mainUrl = 'https://plast.sitegist.net/'
browser.get(mainUrl)
loginButton = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'loginLink')))
loginButton.click()
browser.implicitly_wait(10)
loginField = browser.find_element_by_name('f_login_email')
loginField.send_keys('slavko.senkiv@gmail.com')
loginPass = browser.find_element_by_name('f_login_password')
loginPass.send_keys('makeself')
enterButton = browser.find_element_by_xpath('/html/body/div/div[1]/div/div[3]/div[3]/input')
enterButton.click()

# navigating to TA
time.sleep(2)
baseButton = browser.find_element_by_class_name('handbook')
baseButton.click()
usersButton = browser.find_element_by_link_text('Користувачі')
usersButton.click()
maleCheck = browser.find_element_by_name('sex-male')
maleCheck.click()
stanucjaField = browser.find_element_by_name('stanucja')
stanucjaField.send_keys('Львів')
time.sleep(1)
lvivButton = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]/div[4]/span[1]')
lvivButton.click()
searchButton = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[8]/input')
searchButton.click()

wb = openpyxl.open('епласт Львів хлопці.xlsx')
sheet = wb.active

# scrapping TA data
users_ids_list = []
usersList = browser.find_elements_by_class_name('user')
for user in usersList:
    user_id = user.get_attribute('joinobjid')
    users_ids_list.append(user_id)

queue = 1
# for id in range(5):
for id in users_ids_list:

    print(f'{queue} / {str(len(users_ids_list))}')
    queue += 1

    # user_id
    sheet.cell(row=id + 2, column=1).value = users_ids_list[id]
    user_link = 'https://plast.sitegist.net/?pageid=500&userid=' + users_ids_list[id]
    browser.get(user_link)

    # user_name
    user_name = browser.find_element_by_class_name('profileName').text
    sheet.cell(row=id + 2, column=2).value = user_name

    # user_social
    try:
        user_social = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/a').get_attribute('href')
    except NoSuchElementException:
        user_social = 'no social'
    sheet.cell(row=id + 2, column=11).value = user_social

    # user_kureni
    user_kureni = browser.find_elements_by_class_name('simpleunit')
    user_kureni_string = ''
    for kurin in user_kureni:
        user_kureni_string += ', ' + kurin.text
    sheet.cell(row=id + 2, column=4).value = user_kureni_string

    user_stupin = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[4]/div[2]').text
    sheet.cell(row=id + 2, column=5).value = user_stupin

    user_upu_stupin = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[15]/div[2]').text
    sheet.cell(row=id + 2, column=6).value = user_upu_stupin

    # user_awards
    user_awards = browser.find_elements_by_class_name('award_item')
    user_awards_string = ''
    for award in user_awards:
        user_awards_string += ', ' + award.text
    sheet.cell(row=id + 2, column=10).value = user_awards_string

    # kv
    user_kvs = browser.find_elements_by_class_name('kv_item')
    user_kvs_string = ''
    for kv in user_kvs:
        user_kvs_string += ', ' + kv.get_attribute('title')
    sheet.cell(row=id + 2, column=9).value = user_kvs_string

    user_education = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[10]/div[2]').text
    sheet.cell(row=id + 2, column=12).value = user_education

    #membership
    membershipButton = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[1]/ul/li[2]/a')
    membershipButton.click()

    # startdate
    try:
        startdate = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]/div[1]/div[2]').text
    except NoSuchElementException:
        startdate = 'no startdate'
    sheet.cell(row=id + 2, column=3).value = startdate

    # oathdate
    try:
        oathdate = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]/div[2]/div[2]').text
    except NoSuchElementException:
        oathdate = 'no oathdate'
    sheet.cell(row=id + 2, column=7).value = oathdate

    # stpl_stupin_date
    try:
        stpl_stupin_date = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/table/tbody/tr/td[3]').text
    except NoSuchElementException:
        stpl_stupin_date = 'no stpl_stupin_date'
    sheet.cell(row=id + 2, column=8).value = stpl_stupin_date

wb.save('епласт Львів хлопці.xlsx')















